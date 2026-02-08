"""
Red Bull Network Optimizer - Flask API
Serves both the web interface and RESTful API endpoints.
"""

from flask import Flask, render_template, jsonify, request, send_file
from flask_cors import CORS
import sys
import os

# Add project root to path
sys.path.insert(0, os.path.dirname(__file__))

from optimization.scenario_engine import ScenarioEngine
from config import Config

app = Flask(__name__)
app.config.from_object(Config)
CORS(app)

# Initialize scenario engine
scenario_engine = ScenarioEngine(data_dir='data')

# Cache for storing results
results_cache = {}


@app.route('/')
def index():
    """Serve the main application page."""
    return render_template('index.html')


@app.route('/api/scenarios', methods=['GET'])
def get_scenarios():
    """
    Get list of available optimization scenarios.
    
    Returns:
        JSON with scenario definitions including objectives and use cases
    """
    try:
        scenarios = scenario_engine.get_scenario_definitions()
        return jsonify({
            'success': True,
            'scenarios': scenarios
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/optimize', methods=['POST'])
def run_optimization():
    """
    Run network optimization for selected scenario.
    
    Request body:
        {
            "scenario_id": "baseline" | "cost_optimized" | "disruption"
        }
    
    Returns:
        JSON with solution, KPIs, insights, and comparison to baseline
    """
    try:
        data = request.get_json()
        scenario_id = data.get('scenario_id', 'baseline')
        
        # Validate scenario
        valid_scenarios = ['baseline', 'cost_optimized', 'disruption']
        if scenario_id not in valid_scenarios:
            return jsonify({
                'success': False,
                'error': f'Invalid scenario: {scenario_id}. Must be one of {valid_scenarios}'
            }), 400
        
        # Run optimization
        result = scenario_engine.run_scenario(scenario_id)
        
        # Cache result
        results_cache[scenario_id] = result
        
        # Format response (convert numpy types to native Python)
        response = {
            'success': True,
            'scenario_id': result['scenario_id'],
            'kpis': result['kpis'],
            'insights': result['insights'],
            'comparison': result['comparison'],
            'solution_summary': {
                'total_cost': float(result['solution']['objective_value']),
                'status': result['solution']['status'],
                'production_by_plant': {
                    k: float(v) for k, v in result['solution']['production'].items()
                },
                'cost_breakdown': {
                    k: float(v) for k, v in result['solution']['cost_breakdown'].items()
                }
            }
        }
        
        return jsonify(response)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/kpis', methods=['GET'])
def get_kpis():
    """
    Get current KPIs (from baseline or last run scenario).
    
    Returns:
        JSON with KPI values and business context
    """
    try:
        # Get baseline KPIs if not already cached
        if 'baseline' not in results_cache:
            result = scenario_engine.run_scenario('baseline')
            results_cache['baseline'] = result
        
        baseline = results_cache['baseline']
        
        return jsonify({
            'success': True,
            'kpis': baseline['kpis']
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/network-data', methods=['GET'])
def get_network_data():
    """
    Get network topology data for visualization.
    
    Returns:
        JSON with plants, DCs, markets, and flows
    """
    try:
        import pandas as pd
        
        # Load network data
        plants = pd.read_csv('data/plants.csv')
        dcs = pd.read_csv('data/distribution_centers.csv')
        markets = pd.read_csv('data/markets.csv')
        routes = pd.read_csv('data/transportation.csv')
        
        # Get baseline flows if available
        flows = {}
        if 'baseline' in results_cache:
            baseline = results_cache['baseline']
            flows = {
                'plant_to_dc': [
                    {
                        'from': k[0],
                        'to': k[1],
                        'volume': float(v),
                        'type': 'plant_to_dc'
                    }
                    for k, v in baseline['solution']['plant_to_dc_flows'].items()
                    if v > 1000  # Only significant flows
                ],
                'dc_to_market': [
                    {
                        'from': k[0],
                        'to': k[1],
                        'volume': float(v),
                        'type': 'dc_to_market'
                    }
                    for k, v in baseline['solution']['dc_to_market_flows'].items()
                    if v > 1000
                ]
            }
        
        return jsonify({
            'success': True,
            'plants': plants.to_dict('records'),
            'dcs': dcs.to_dict('records'),
            'markets': markets.to_dict('records'),
            'flows': flows
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """
    Generate Excel export with scenario results.
    
    Request body:
        {
            "scenario_id": "baseline" | "cost_optimized" | "disruption"
        }
    
    Returns:
        Excel file download
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        
        data = request.get_json()
        scenario_id = data.get('scenario_id', 'baseline')
        
        # Get scenario result
        if scenario_id not in results_cache:
            result = scenario_engine.run_scenario(scenario_id)
            results_cache[scenario_id] = result
        else:
            result = results_cache[scenario_id]
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Header
        ws['A1'] = "Red Bull Global Network Optimizer"
        ws['A1'].font = Font(bold=True, size=16, color="DB0A40")
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Scenario: {scenario_id.upper()}"
        ws['A2'].font = Font(bold=True, size=12)
        
        # KPI table
        ws['A4'] = "Key Performance Indicators"
        ws['A4'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ['Metric', 'Value', 'vs Target', 'Business Impact']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # KPI data
        kpis = result['kpis']
        kpi_rows = [
            ['Total Network Cost', kpis['total_cost']['formatted'], kpis['total_cost']['vs_optimal'], kpis['total_cost']['impact']],
            ['Fill Rate', kpis['fill_rate']['formatted'], kpis['fill_rate']['vs_target'], kpis['fill_rate']['impact']],
            ['Avg Lead Time', kpis['avg_lead_time']['formatted'], kpis['avg_lead_time']['vs_competitor'], kpis['avg_lead_time']['impact']],
            ['CO2 Emissions', kpis['co2_emissions']['formatted'], kpis['co2_emissions']['vs_target'], kpis['co2_emissions']['impact']],
            ['Cost per Unit', kpis['cost_per_unit']['formatted'], kpis['cost_per_unit']['vs_budget'], kpis['cost_per_unit']['impact']],
        ]
        
        for row_idx, row_data in enumerate(kpi_rows, start=6):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 60
        
        # Insights sheet
        ws2 = wb.create_sheet("Strategic Insights")
        ws2['A1'] = "Strategic Recommendations"
        ws2['A1'].font = Font(bold=True, size=14)
        
        row = 3
        for insight in result['insights']:
            ws2[f'A{row}'] = insight['title']
            ws2[f'A{row}'].font = Font(bold=True, size=12)
            ws2[f'B{row}'] = f"Priority: {insight['priority'].upper()}"
            
            ws2[f'A{row+1}'] = insight['description']
            ws2[f'A{row+2}'] = f"Impact: {insight['impact']}"
            ws2[f'A{row+3}'] = f"Implementation: {insight['implementation']}"
            
            row += 5
        
        ws2.column_dimensions['A'].width = 80
        ws2.column_dimensions['B'].width = 20
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'red_bull_network_analysis_{scenario_id}.xlsx'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'version': '1.0.0'
    })


if __name__ == '__main__':
    print("🚀 Starting Red Bull Network Optimizer...")
    print("📍 Server running at: http://localhost:5000")
    print("📊 API endpoints available at: http://localhost:5000/api/")
    print("\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
